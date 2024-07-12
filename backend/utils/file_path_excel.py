from pathlib import Path
from openpyxl import load_workbook

from openpyxl.styles import Alignment

def file_path_excel(cam_name):
    base_dir = Path(__file__).resolve().parent.parent  
    directory = base_dir / 'result'
    excel_file = f"{cam_name}_result.xlsx"
    file_path = directory / excel_file
    return file_path
    
    
def read_excel_to_json(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    headers = [cell.value.strip() if cell.value else "null" for cell in sheet[2]]
    data = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        row_data = {headers[i]: row[i] for i in range(len(headers))}
        data.append(row_data)
    
    return data


    
def update_excel_file(file_path):

    workbook = load_workbook(file_path)
    sheet = workbook.active
    center_alignment = Alignment(horizontal='center', vertical='center')
    for row in sheet.iter_rows(min_row=3, min_col=1, max_col=8):  
        for cell in row[4:7]:  
            if cell.value is None or cell.value == "":
                cell.value = "✗"
            cell.alignment = center_alignment
    workbook.save(file_path)