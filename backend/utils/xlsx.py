
from openpyxl import Workbook , load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_excel_file(file_path, targets, owner_email, name_of_campaign):

    wb = Workbook()
    ws = wb.active

    ws.title = "Result campaign"

    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"{name_of_campaign} campaign of {owner_email}"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    headers = ["Amount", "Firstname", "Lastname", "Email", "Open email", "Click link", "Submit data", "Error"]
    header_fill = PatternFill(start_color="FFCC00", end_color="FFCC00", fill_type="solid")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill

    for row_num, target in enumerate(targets, 3):
        amount_cell = ws.cell(row=row_num, column=1)
        amount_cell.value = row_num - 2  
        amount_cell.alignment = Alignment(horizontal='center')
        
        ws.cell(row=row_num, column=2).value = target.get('first_name')
        ws.cell(row=row_num, column=3).value = target.get('last_name')
        ws.cell(row=row_num, column=4).value = target.get('email')
        ws.cell(row=row_num, column=5).value = ""  # Open email
        ws.cell(row=row_num, column=6).value = ""  # Click link
        ws.cell(row=row_num, column=7).value = ""  # Submit data
        ws.cell(row=row_num, column=8).value = ""  # Error

    # Adjust column widths
    column_widths = [10, 15, 15, 25, 15, 15, 15, 10]
    for i, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = column_width

    # Save the workbook
    wb.save(file_path)



def edit_excel_file(file_path, email, column_index, char="✗"):
    wb = load_workbook(file_path)
    sheet = wb.active
    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=4, max_col=4):
        for cell in row:
            # Check if the email matches and edit column 8 if true
            if cell.value == email:
                row_index = cell.row
                cell_data = sheet.cell(row=row_index, column = column_index)
                cell_data.value = char
                cell_data.alignment = Alignment(horizontal='center', vertical='center')
                break

    wb.save(file_path)
    

def check_empty(file_path, email, column_index):
    wb = load_workbook(file_path)
    sheet = wb.active

    for row in range(2, sheet.max_row + 1):
        email_value = sheet.cell(row=row, column=4).value  # Assuming the 'Email' column is at index 4 (D)
        if email_value == email:
            open_email_value = sheet.cell(row=row, column=column_index).value
            return open_email_value is not None and open_email_value != ""

    return False
